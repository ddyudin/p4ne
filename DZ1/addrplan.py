import glob
import re
import ipaddress
from openpyxl import * #найти функции создания книги, листа, записи на лист

list_of_address = set()
for file_name in glob.glob(('f:\\Seafile\\p4ne_training\\config_files\\*.txt')):
    with open(file_name) as f:
        for imp_str in f:
            s = re.match('^ ip address ([0-9]+\.[0-9]+\.[0-9]+\.[0-9]+) ([0-9]+\.[0-9]+\.[0-9]+\.[0-9]+)', imp_str)
            if s:
                address = (s.group(1), s.group(2))
                list_of_address.add(address)

list_of_networks = set()
for address in list_of_address:
    list_of_networks.add(ipaddress.ip_network(address[0] + '/' + address[1], strict = False))
ln2 = sorted(list_of_networks)

wb = Workbook()
ws = wb.active

ws['A1'] = 'Network'
ws['B1'] = 'Mask'
ws['C1'] = 'Gateway'
current_row = 2

for network in ln2:
    for address in list_of_address:
        if ipaddress.IPv4Interface(address[0] + '/' + address[1]) in network:
            print('%s\t\t\t%s\t\t%s' % (str(network), address[1], address[0]))
            ws.cell(row=current_row, column=1, value=str(network))
            ws.cell(row=current_row, column=2, value=address[1])
            ws.cell(row=current_row, column=3, value=address[0])
            current_row += 1
            continue

wb.save('addrplan.xlsx')